import os
from pathlib import Path
import sys
import fire
import pandas as pd
from lxml import etree
from flatten_dict import flatten, unflatten
from openpyxl.styles import Alignment
import module

def xml_to_excel(elem_table: str, xmldir: str, outdir: str='./'):
    """
    Convert multiple XML files into a unified Excel table.
    
    Parameters
    ----------
    elem_table: str
        Excel table defining element names in the output Excel data each corresponding to the XPath of the input XML file.
    xmldir: str
        Path to the folder containing input XML files.
    outdir: str
        Output directory. Default is './'.
    """

    print('### START ###')
    print('element define: ' + elem_table)
    print('xmldir: ' + xmldir)
    print('outdir: ' + outdir)

    if not os.path.exists(outdir):
        os.mkdir(outdir)

    xmlfiles = [str(p) for p in Path(xmldir).rglob('*.xml')]
    print('num of xmlfiles: ' + str(len(xmlfiles)))

    print('')
    print('Read Element-Define table')

    dict_path = module.read_define(elem_table, 'XPath')

    table_path = [l for l in module.dict_to_table(dict_path, 'XPath')]
    table_name_define = [l[0] for l in table_path]
    data_path = pd.DataFrame(table_name_define)
    cols_name = list(data_path.columns)
    data_path['XPath'] = [l[1] for l in table_path]
    data_path = module.data_to_str(data_path)

    outfile = outdir + '/table_path.xlsx'
    data_path.to_excel(outfile, index=False)
    print('output: ' + outfile)

    print('')
    print('Compile XML values')

    n = 0
    pref_default = '_ns'
    for xmlfile in xmlfiles:
        xmlfile0 = xmlfile.split('/')[-1]
        print(xmlfile0, end=' ', flush=True)

        xml = etree.parse(xmlfile)
        root = xml.getroot()
        ns_tmp = root.nsmap
        ns = {}
        for key_tmp in ns_tmp:
            key = pref_default if not key_tmp else key_tmp
            ns[key] = ns_tmp[key_tmp]

        table_xml = [l for l in module.read_xml(dict_path, root, ns, pref_default, 'XPath')]

        table_name = [l[0] for l in table_xml]
        data_xml = pd.DataFrame(table_name)
        for col in cols_name:
            if col not in data_xml.columns:
                data_xml[col] = None
        data_xml = module.data_to_str(data_xml)

        data_xml[xmlfile0] = [l[1] for l in table_xml]
        if n>0:
            data_merge = pd.merge(data_merge, data_xml, on = cols_name, how = 'outer')
        else:
            data_merge = data_xml

        n += 1
        if n%1000 == 0:
            print('')
            print('#' + str(n))

    print('')
    data_merge = module.data_to_str(data_merge)

    print('')
    print('Rearrange row order')
    dict_flatten = {}
    for i, row in data_merge.iterrows():
        dict_flatten['/'.join([s for s in row[cols_name] if s!=''])] = row[[s for s in row.index if not s in cols_name]].tolist()
    dict_data = unflatten(dict_flatten, splitter='path')
    dict_data_sort = module.reorder_dict(dict_path, dict_data)
    dict_flatten = flatten(dict_data_sort)

    table_name = [[s for s in t] for t in dict_flatten.keys()]
    data_name = pd.DataFrame(table_name)
    for col in cols_name:
        if col not in data_name.columns:
            data_name[col] = None

    list_val = [l for l in dict_flatten.values()]
    data = pd.concat([data_name, pd.DataFrame(list_val)], axis=1)
    data.columns = data_merge.columns
    data = module.data_to_str(data)
    cols_name_new = []
    for col in cols_name:
        if (data[col]=='').all():
            data = data.drop(columns=col)
            continue
        cols_name_new.append(col)

    print('')
    print('Output')
    outfile = outdir + '/table.pkl'
    module.pickle_dump(data, outfile)
    print(outfile)

    outfile = outdir + '/table.xlsx'
    if os.path.exists(outfile):
        os.remove(outfile)

    with pd.ExcelWriter(outfile, engine='openpyxl') as writer:
        data.to_excel(writer, sheet_name='Sheet1', index=False)
        ws = writer.sheets['Sheet1']

        print('Merge Excel cells')
        start_row = 2
        end_row = start_row + len(data) - 1
        dict_merge_v = {}
        for col in range(1, len(cols_name_new) + 1):
            dict_merge_v[col] = module.list_merge_vertical(ws, col, start_row, end_row)

        dict_merge_h = module.dict_merge_horizontal(ws, start_row, end_row, 1, len(cols_name_new))

        for col, list_merge in dict_merge_v.items():
            for lim in list_merge:
                ws.merge_cells(start_row = lim[0], start_column = col, end_row = lim[1], end_column = col)

        for row, list_merge in dict_merge_h.items():
            if len(list_merge) > 1:
                ws.merge_cells(start_row = row, start_column = list_merge[0], end_row = row, end_column = list_merge[1])
    
        print('Remove temporary suffix in element names')
        for row in range(start_row, end_row + 1):
            for col in range(1, len(cols_name_new) + 1):
                cell = ws.cell(row = row, column = col)
                cell.alignment = Alignment(horizontal = 'left', vertical = 'center')
                if isinstance(cell.value, str):
                    cell.value = cell.value.rsplit('_', 1)[0]

    print('output: ' + outfile)

    print('Finished!')

if __name__ == "__main__":
    fire.Fire(xml_to_excel)
